import os
from dataclasses import dataclass
from typing import List, Sequence, Tuple

import numpy as np
import torch
from PIL import Image
from torch.utils.data import Dataset
from torchvision import transforms
from torchvision.transforms import functional as TF


def read_score_table(score_xlsx: str) -> List[Tuple[str, float]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Train_scores.xlsx") from exc

    wb = load_workbook(score_xlsx, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty score file: {score_xlsx}")

    data_rows = rows[1:] if isinstance(rows[0][0], str) else rows
    pairs: List[Tuple[str, float]] = []
    for row in data_rows:
        if row is None or len(row) < 2:
            continue
        raw_name = row[0]
        raw_score = row[1]
        if raw_name is None or raw_score is None:
            continue
        name = str(raw_name).strip()
        if not name.lower().endswith(".png"):
            name = f"{name}.png"
        score = float(raw_score)
        pairs.append((name, score))

    if not pairs:
        raise ValueError(f"No valid rows parsed from: {score_xlsx}")
    return pairs


class ResizePadSquare:
    def __init__(self, size: int):
        self.size = size

    def __call__(self, img: Image.Image) -> Image.Image:
        w, h = img.size
        scale = self.size / max(w, h)
        nw = max(1, int(round(w * scale)))
        nh = max(1, int(round(h * scale)))
        img = img.resize((nw, nh), Image.BICUBIC)

        pad_w = self.size - nw
        pad_h = self.size - nh
        left = pad_w // 2
        right = pad_w - left
        top = pad_h // 2
        bottom = pad_h - top
        return TF.pad(img, [left, top, right, bottom], fill=0)


def build_train_transform(image_size: int, mean=None, std=None) -> transforms.Compose:
    mean = mean or [0.485, 0.456, 0.406]
    std = std or [0.229, 0.224, 0.225]
    return transforms.Compose(
        [
            ResizePadSquare(image_size),
            transforms.RandomHorizontalFlip(p=0.5),
            transforms.ToTensor(),
            transforms.Normalize(mean=mean, std=std),
        ]
    )


def build_eval_transform(image_size: int, mean=None, std=None) -> transforms.Compose:
    mean = mean or [0.485, 0.456, 0.406]
    std = std or [0.229, 0.224, 0.225]
    return transforms.Compose(
        [
            ResizePadSquare(image_size),
            transforms.ToTensor(),
            transforms.Normalize(mean=mean, std=std),
        ]
    )


def stratified_split_indices(scores: Sequence[float], val_ratio: float, seed: int) -> Tuple[List[int], List[int]]:
    bins = {}
    for idx, score in enumerate(scores):
        key = int(round(score))
        bins.setdefault(key, []).append(idx)

    rng = np.random.default_rng(seed)
    train_ids: List[int] = []
    val_ids: List[int] = []

    for group in bins.values():
        group = np.array(group)
        rng.shuffle(group)
        n_val = max(1, int(round(len(group) * val_ratio)))
        val_ids.extend(group[:n_val].tolist())
        train_ids.extend(group[n_val:].tolist())

    return sorted(train_ids), sorted(val_ids)


@dataclass
class PairSample:
    name: str
    score: float


class SemanticIQADataset(Dataset):
    def __init__(self, samples: List[PairSample], ref_dir: str, dist_dir: str, transform=None):
        self.samples = samples
        self.ref_dir = ref_dir
        self.dist_dir = dist_dir
        self.transform = transform

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, idx: int):
        sample = self.samples[idx]
        ref_path = os.path.join(self.ref_dir, sample.name)
        dist_path = os.path.join(self.dist_dir, sample.name)

        if not os.path.exists(ref_path):
            raise FileNotFoundError(f"Missing ref image: {ref_path}")
        if not os.path.exists(dist_path):
            raise FileNotFoundError(f"Missing dist image: {dist_path}")

        ref_img = Image.open(ref_path).convert("RGB")
        dist_img = Image.open(dist_path).convert("RGB")

        if self.transform is not None:
            ref_img = self.transform(ref_img)
            dist_img = self.transform(dist_img)

        score_float = torch.tensor(sample.score, dtype=torch.float32)
        score_cls = torch.tensor(int(round(sample.score)), dtype=torch.long)
        return {
            "name": sample.name,
            "ref": ref_img,
            "dist": dist_img,
            "score": score_float,
            "score_cls": score_cls,
        }
